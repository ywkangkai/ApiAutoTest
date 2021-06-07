import openpyxl
import os
import common.logHandler as myLogging


class XlsHandlerConf:
    def __init__(self):
        global myLogger
        myLogger = myLogging.Logger()

    # 打开指定的.xlsx文件，通过一个二维list传入Excel表中，多次调用将覆盖
    # 示例：params = [[l,2,3],[4,5,6]] 将1,2,3依次写入第一行，4,5,6依次写入第二行
    def write_data_in_workbook(self, filename, params):
        try:
            file_path = os.path.join(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data'), filename)
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            for i in range(len(params)):
                for j in range(len(params[i])):
                    worksheet.cell(i + 1, j + 1, params[i][j])
            workbook.save(file_path)
        except Exception as msg:
            myLogger.logger.error(msg)

    # 将写入的数据读取想要的value进行断言操作
    # 获取写入文件中所有的值，并反馈一个list
    # 用例中根据值是否在list中判定是否成功导入
    def read_data_in_workbook(self, filename):
        try:
            file_path = os.path.join(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data'), filename)
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            data_list = []
            for i in range(1, worksheet.max_row + 1):
                for j in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=i, column=j).value
                    data_list.append(cell_value)
            return data_list
        except Exception as msg:
            myLogger.logger.error(msg)

    # 将写入的数据清空，保存为一个空文件
    def clear_data_in_workbook(self, filename):
        try:
            file_path = os.path.join(os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data'), filename)
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            for i in range(1, worksheet.max_row + 1):
                for j in range(1, worksheet.max_column + 1):
                    worksheet.cell(row=i, column=j, value="")
            workbook.save(file_path)
        except Exception as msg:
            myLogger.logger.error(msg)
